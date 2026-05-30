#!/usr/bin/env python3
"""
Aegis Omniscient v12.0 – Professional Stock Valuation Terminal
- Multi‑stage DCF with smart growth (analyst, historical, sustainable)
- Implied market growth (reverse DCF)
- Analyst ratings & price targets
- Monte Carlo simulation (optional)
- Peer percentile rankings
- Full export (Excel, PDF, JSON, CSV)
- No API keys – 100% free
"""

import typer
import yfinance as yf
import numpy as np
import pandas as pd
import json
import os
import math
from datetime import datetime
from functools import lru_cache
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.columns import Columns
from rich.prompt import Prompt
from rich.progress import Progress, SpinnerColumn, TextColumn
import warnings
warnings.filterwarnings("ignore")

# Optional imports for export & scraping
try:
    from fpdf import FPDF
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = typer.Typer(help="🧠 Aegis Omniscient v12.0 – Institutional Valuation")
console = Console()

# ---------- CONFIG ----------
CONFIG_FILE = "aegis_config.json"
WATCHLIST_FILE = "aegis_watchlist.json"
REPORT_DIR = "aegis_reports"
CACHE_DIR = os.path.join(REPORT_DIR, "cache")
os.makedirs(REPORT_DIR, exist_ok=True)
os.makedirs(CACHE_DIR, exist_ok=True)

DEFAULT_CONFIG = {
    "default_ticker": "AAPL",
    "use_monte_carlo": False,
    "num_simulations": 5000,
    "max_growth": 0.20,
    "max_stage1_growth": 0.20,
    "min_growth": 0.01,
    "wacc_min": 0.045,
    "wacc_max": 0.15,
    "equity_risk_premium": 0.05,
    "default_base": "ocf",
    "add_net_cash": True,
    "confidence_weights": {
        "mos": 0.35,
        "analyst": 0.20,
        "technical": 0.20,
        "quality": 0.25
    }
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            return {**DEFAULT_CONFIG, **json.load(f)}
    return DEFAULT_CONFIG.copy()

def save_config(config):
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=2)

# ---------- CACHE HELPERS ----------
def cache_get(key, expiry_hours=6):
    path = os.path.join(CACHE_DIR, f"{key}.json")
    if os.path.exists(path):
        with open(path, "r") as f:
            data = json.load(f)
        if datetime.now().timestamp() - data['ts'] < expiry_hours * 3600:
            return data['value']
    return None

def cache_set(key, value):
    path = os.path.join(CACHE_DIR, f"{key}.json")
    with open(path, "w") as f:
        json.dump({"ts": datetime.now().timestamp(), "value": to_json_safe(value)}, f)

# ---------- JSON-SAFE SERIALIZATION ----------
def to_json_safe(obj):
    """Recursively convert numpy / pandas scalars and containers into plain
    Python types so results can be cached and exported as JSON."""
    if isinstance(obj, dict):
        return {str(k): to_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [to_json_safe(v) for v in obj]
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        val = float(obj)
        return None if math.isnan(val) else val
    if isinstance(obj, np.ndarray):
        return [to_json_safe(v) for v in obj.tolist()]
    if isinstance(obj, float):
        return None if math.isnan(obj) else obj
    return obj

# ---------- SAFE SCALAR ----------
def safe_scalar(value, default=0.0):
    if value is None:
        return float(default)
    if isinstance(value, (pd.Series, pd.DataFrame)):
        if len(value) == 0:
            return float(default)
        value = value.iloc[0]
    if isinstance(value, (np.ndarray, list, tuple)):
        if len(value) == 0:
            return float(default)
        value = value[0]
    try:
        val = float(value)
        if np.isnan(val):
            return float(default)
        return val
    except (TypeError, ValueError):
        return float(default)

def find_row(df, possible_names, default=0.0):
    if df is None or df.empty:
        return safe_scalar(default)
    for name in possible_names:
        if name in df.index:
            return safe_scalar(df.loc[name])
    for idx in df.index:
        for pname in possible_names:
            if pname.lower() in idx.lower():
                return safe_scalar(df.loc[idx])
    return safe_scalar(default)

def normalized_fcf_per_share(cashflow, shares, years=3):
    """Average of the last `years` of free cash flow (OCF + capex), per share.

    A single year's FCF is noisy — capex spikes (e.g. a data-center buildout)
    can crush one year and wreck the whole DCF. Averaging the most recent few
    years gives a far more stable base. Returns (fcf_per_share, n_years_used).
    """
    if cashflow is None or cashflow.empty or not shares:
        return 0.0, 0
    ocf_row = cashflow.loc['Operating Cash Flow'] if 'Operating Cash Flow' in cashflow.index else None
    cap_row = None
    for name in ['Capital Expenditure', 'Capital Expenditures']:
        if name in cashflow.index:
            cap_row = cashflow.loc[name]
            break
    if ocf_row is None:
        return 0.0, 0
    fcfs = []
    for i in range(min(years, len(cashflow.columns))):
        ocf = safe_scalar(ocf_row.iloc[i])
        capex = safe_scalar(cap_row.iloc[i]) if cap_row is not None else 0.0
        if ocf != 0:
            fcfs.append(ocf + capex)
    if not fcfs:
        return 0.0, 0
    return (float(np.mean(fcfs)) / shares), len(fcfs)

# ---------- RISK FREE RATE ----------
@lru_cache(maxsize=1)
def get_risk_free_rate():
    try:
        tnx = yf.Ticker("^TNX").history(period="1d")['Close'].iloc[-1]
        return safe_scalar(tnx) / 100
    except Exception:
        return 0.042

# ---------- MARKET RISK PREMIUM ----------
def get_market_risk_premium():
    # Using historical average 5-6% for US market
    return 0.055

# ---------- CURRENCY NORMALIZATION ----------
# Some venues quote prices in a *minor* unit (1/100 of the major currency) while
# the financial statements are reported in the major unit. The classic trap is
# London: a `.L` ticker's price currency is "GBp" (pence) but its statements are
# in "GBP" — a silent 100x mismatch that wrecks any per-share comparison.
MINOR_UNITS = {
    "GBp": ("GBP", 100.0),   # pence
    "GBX": ("GBP", 100.0),   # pence (alt code)
    "ZAc": ("ZAR", 100.0),   # SA cents
    "ILA": ("ILS", 100.0),   # Israeli agorot
}

def _major_unit(ccy):
    """Return (major_currency, minor_per_major). For a normal currency this is
    (ccy, 1.0); for a minor unit like GBp it's ('GBP', 100.0)."""
    if not ccy:
        return None, 1.0
    return MINOR_UNITS.get(ccy, (ccy, 1.0))

@lru_cache(maxsize=64)
def get_fx_rate(from_ccy, to_ccy):
    """Spot FX rate to convert 1 unit of `from_ccy` into `to_ccy` (major units).
    Uses yfinance's `<PAIR>=X` quotes; returns 1.0 on failure or when equal."""
    if not from_ccy or not to_ccy or from_ccy == to_ccy:
        return 1.0
    try:
        hist = yf.Ticker(f"{from_ccy}{to_ccy}=X").history(period="5d")["Close"]
        rate = safe_scalar(hist.iloc[-1]) if not hist.empty else 0.0
        if rate > 0:
            return float(rate)
    except Exception:
        pass
    return 1.0

def convert_currency(value, from_ccy, to_ccy):
    """Convert a monetary `value` expressed in `from_ccy` into `to_ccy`,
    correctly handling minor units (GBp/pence etc.) on either side."""
    if value is None:
        return value
    from_major, from_div = _major_unit(from_ccy)
    to_major, to_mult = _major_unit(to_ccy)
    if not from_major or not to_major:
        return value
    rate = get_fx_rate(from_major, to_major)
    # value(from minor) -> from major -> to major -> to minor
    return value / from_div * rate * to_mult

# ---------- SMART GROWTH ESTIMATION ----------
def estimate_growth(ticker, financials, cashflow, info, shares):
    cfg = load_config()
    growth_candidates = []
    sources = []

    # 1) Analyst long-term growth (earnings_estimate)
    try:
        est = ticker.earnings_estimate
        if est is not None and 'avg' in est.index:
            val = safe_scalar(est.loc['avg'].iloc[0])
            if val > 1:
                val /= 100
            if cfg['min_growth'] <= val <= cfg['max_growth']:
                growth_candidates.append(val)
                sources.append("Analyst LT growth")
    except Exception:
        pass

    # 2) Historical revenue CAGR (last 5 years)
    try:
        revenue_vals = []
        if 'Total Revenue' in financials.index:
            for i in range(min(5, len(financials.columns))):
                rev = safe_scalar(financials.loc['Total Revenue'].iloc[i])
                if rev > 0:
                    revenue_vals.append(rev)
        if len(revenue_vals) >= 3:
            cagr = (revenue_vals[0] / revenue_vals[-1]) ** (1/(len(revenue_vals)-1)) - 1
            if cfg['min_growth'] <= cagr <= cfg['max_growth']:
                growth_candidates.append(cagr)
                sources.append("Historical revenue CAGR")
    except Exception:
        pass

    # 3) Sustainable growth (ROE * retention ratio)
    try:
        net_income = find_row(financials, ['Net Income', 'Net Income Continuous Operations'], 0.0)
        total_equity = find_row(ticker.balance_sheet, ['Total Equity Gross Minority Interest', 'Total Equity'], 1.0)
        if total_equity != 0 and net_income != 0:
            roe = net_income / total_equity
            dividends = find_row(cashflow, ['Dividends Paid', 'Common Stock Dividends'], 0.0)
            payout = dividends / net_income if net_income != 0 else 0
            sustainable = roe * (1 - payout)
            if cfg['min_growth'] <= sustainable <= cfg['max_growth']:
                growth_candidates.append(sustainable)
                sources.append("Sustainable (ROE×retention)")
    except Exception:
        pass

    # 4) Historical FCF/share growth (last 5 years)
    try:
        fcf_vals = []
        for i in range(min(5, len(cashflow.columns))):
            ocf = safe_scalar(cashflow.loc['Operating Cash Flow'].iloc[i]) if 'Operating Cash Flow' in cashflow.index else 0
            capex = safe_scalar(cashflow.loc['Capital Expenditure'].iloc[i]) if 'Capital Expenditure' in cashflow.index else 0
            fcf_vals.append(ocf + capex)
        if len(fcf_vals) >= 3 and fcf_vals[-1] != 0:
            cagr = (fcf_vals[0] / fcf_vals[-1]) ** (1/(len(fcf_vals)-1)) - 1
            if cfg['min_growth'] <= cagr <= cfg['max_growth']:
                growth_candidates.append(cagr)
                sources.append("Historical FCF CAGR")
    except Exception:
        pass

    # Choose the best (median of candidates, or default 8%)
    if growth_candidates:
        growth = np.median(growth_candidates)
        console.print(f"[dim]✓ Growth: {growth:.1%} ({', '.join(sources[:2])})[/dim]")
    else:
        growth = 0.08
        console.print("[dim]✓ Growth: default 8% (sector average)[/dim]")
    return growth

# ---------- COUNTRY RISK PREMIUM ----------
# Damodaran-style equity country-risk premiums (approximate, mid-2024; refresh
# periodically). Developed markets ≈ 0; emerging markets carry a premium for
# sovereign/FX risk. Keyed on yfinance's info['country'] (HQ country).
COUNTRY_RISK_PREMIUM = {
    # developed (no premium)
    "United States": 0.0, "Canada": 0.0, "United Kingdom": 0.0, "Germany": 0.0,
    "France": 0.0, "Switzerland": 0.0, "Netherlands": 0.0, "Sweden": 0.0,
    "Norway": 0.0, "Denmark": 0.0, "Australia": 0.0, "New Zealand": 0.0,
    "Japan": 0.0058, "Singapore": 0.0, "Ireland": 0.0058, "Hong Kong": 0.0057,
    "Taiwan": 0.0057, "South Korea": 0.0086,
    # emerging
    "China": 0.0086, "India": 0.0231, "Brazil": 0.0345, "Mexico": 0.0173,
    "Uruguay": 0.0173, "Chile": 0.0115, "Colombia": 0.0288, "Peru": 0.0173,
    "Argentina": 0.1265, "South Africa": 0.0345, "Indonesia": 0.0173,
    "Philippines": 0.0173, "Thailand": 0.0144, "Malaysia": 0.0115,
    "Turkey": 0.0691, "Russia": 0.0750, "Nigeria": 0.0750, "Egypt": 0.0750,
    "Saudi Arabia": 0.0086, "United Arab Emirates": 0.0058, "Israel": 0.0086,
    "Poland": 0.0115, "Greece": 0.0345, "Italy": 0.0173, "Spain": 0.0115,
    "Portugal": 0.0144, "Vietnam": 0.0288,
}

def country_risk_premium(country):
    """Equity country-risk premium for a company's HQ country (0 if unknown
    or developed)."""
    if not country:
        return 0.0
    return COUNTRY_RISK_PREMIUM.get(country.strip(), 0.0)

# ---------- COST OF EQUITY (CAPM) ----------
def cost_of_equity(beta, rf, erp, crp=0.0):
    """CAPM cost of equity with a Blume-adjusted, bounded beta (the discount
    rate used by Simply-Wall-St-style FCFE models), plus an additive country
    risk premium `crp` for sovereign/FX risk in emerging markets."""
    adj_beta = 0.67 * beta + 0.33          # Blume adjustment toward the market
    adj_beta = max(0.8, min(2.0, adj_beta))
    return rf + adj_beta * erp + crp

# ---------- LONG-RUN (TERMINAL) GROWTH ----------
@lru_cache(maxsize=1)
def get_longrun_growth():
    """Perpetual growth rate = 5-year average of the 10-yr Treasury yield
    (Simply Wall St convention), capped to a sane band."""
    try:
        hist = yf.Ticker("^TNX").history(period="5y")['Close']
        avg = safe_scalar(hist.mean()) / 100
        if avg > 0:
            return float(min(0.04, max(0.015, avg)))
    except Exception:
        pass
    return 0.025

# ---------- ANALYST GROWTH SEED ----------
def analyst_growth_rate(ticker, fallback):
    """Near-term growth seed from analyst earnings estimates (avg of current &
    next fiscal year). Falls back to the historical/heuristic estimate."""
    try:
        est = ticker.earnings_estimate
        gs = []
        for period in ('0y', '+1y'):
            if period in est.index:
                g = safe_scalar(est.loc[period, 'growth'])
                if -0.5 < g < 1.0:
                    gs.append(g)
        if gs:
            return float(np.mean(gs))
    except Exception:
        pass
    return fallback

# ---------- 2-STAGE FCFE DCF (Simply Wall St style) ----------
def fcfe_two_stage_parts(base_fcfe, g_start, r, g_term, years=10):
    """Split the 2-stage FCFE value into (PV of explicit cash flows, PV of the
    terminal value). Returns (0.0, 0.0) when the model is undefined."""
    if base_fcfe <= 0 or r <= g_term:
        return 0.0, 0.0
    cfs = []
    tmp = base_fcfe
    for t in range(1, years + 1):
        g = g_start - (g_start - g_term) * ((t - 1) / (years - 1))
        tmp *= (1 + g)
        cfs.append(tmp)
    pv = sum(c / ((1 + r) ** i) for i, c in enumerate(cfs, 1))
    tv = cfs[-1] * (1 + g_term) / (r - g_term)
    tv_pv = tv / ((1 + r) ** years)
    return pv, tv_pv

def fcfe_two_stage(base_fcfe, g_start, r, g_term, years=10):
    """2-stage Free-Cash-Flow-to-Equity DCF.

    Projects `years` of levered FCF with the growth rate fading linearly from
    `g_start` to the long-run rate `g_term`, then a Gordon-Growth terminal
    value, all discounted at the cost of equity `r`. Returns the equity value
    in the same units as `base_fcfe` (per-share in, per-share out)."""
    pv, tv_pv = fcfe_two_stage_parts(base_fcfe, g_start, r, g_term, years)
    return pv + tv_pv

def normalized_cash_earnings_ps(cashflow, financials, shares, mode='auto'):
    """Per-share base cash flow that the FCFE model grows from. `mode` selects
    how aggressively to treat reinvestment:

    - 'ni'   : latest net income (most conservative — assumes today's heavy
               capex persists, so reported FCF stays depressed).
    - 'fcf'  : 3-yr average free cash flow (OCF + capex).
    - 'ocf'  : latest operating cash flow (most Simply-Wall-St-like — treats
               elevated capex as discretionary growth investment that the
               analysts SWS relies on expect to normalize).
    - 'auto' : the greater of 3-yr-avg FCF and net income (default; balanced).
    """
    nfcf, _ = normalized_fcf_per_share(cashflow, shares, years=3)
    ni_ps = find_row(financials, ['Net Income', 'Net Income Continuous Operations'], 0.0) / shares
    ocf_ps = find_row(cashflow, ['Operating Cash Flow', 'Total Cash From Operating Activities'], 0.0) / shares
    if mode == 'ni':
        return ni_ps
    if mode == 'fcf':
        return nfcf
    if mode == 'ocf':
        return ocf_ps if ocf_ps > 0 else (nfcf or ni_ps)
    candidates = [x for x in (nfcf, ni_ps) if x and x > 0]
    if candidates:
        return max(candidates)
    return nfcf if nfcf else ni_ps

# ---------- REVENUE → MARGIN MODEL (pre-profit companies) ----------
# Mature, steady-state net margin and exit P/E by sector. Pre-profit growth
# names (lidar, space, SaaS) have no usable cash flow yet, so we value the
# *future* profitable business: grow revenue, ramp to a sector-normal margin,
# then capitalize. Figures are deliberately conservative peer medians.
SECTOR_MARGIN_EXIT = {
    "Technology":             (0.15, 25.0),
    "Communication Services": (0.15, 22.0),
    "Healthcare":             (0.14, 22.0),
    "Consumer Cyclical":      (0.08, 20.0),
    "Consumer Defensive":     (0.09, 20.0),
    "Industrials":            (0.10, 22.0),
    "Energy":                 (0.10, 12.0),
    "Basic Materials":        (0.10, 14.0),
    "Real Estate":            (0.20, 18.0),
    "Utilities":              (0.10, 16.0),
}
DEFAULT_MARGIN_EXIT = (0.10, 20.0)

def sector_margin_exit(sector):
    """(target steady-state net margin, exit P/E) for a sector."""
    return SECTOR_MARGIN_EXIT.get(sector, DEFAULT_MARGIN_EXIT)

def revenue_growth_seed(info, financials, cap_hi=0.40, floor=0.05):
    """Stage-1 revenue growth for the revenue→margin model: blend of the latest
    YoY revenue growth and the historical revenue CAGR, clamped to a sane band."""
    cands = []
    rg = info.get('revenueGrowth')
    if rg and rg > 0:
        cands.append(float(rg))
    try:
        if financials is not None and 'Total Revenue' in financials.index:
            vals = []
            for i in range(min(5, len(financials.columns))):
                v = safe_scalar(financials.loc['Total Revenue'].iloc[i])
                if v > 0:
                    vals.append(v)
            if len(vals) >= 3:
                cagr = (vals[0] / vals[-1]) ** (1 / (len(vals) - 1)) - 1
                if cagr > 0:
                    cands.append(cagr)
    except Exception:
        pass
    g = float(np.median(cands)) if cands else 0.15
    return max(floor, min(cap_hi, g))

def annual_dilution(cashflow, financials, shares, price_fin):
    """Approximate yearly share-count dilution from stock-based comp: SBC per
    share / price. Pre-profit names pay staff heavily in stock, so ignoring this
    overstates per-share value. Bounded to [0, 6%]."""
    if not shares or not price_fin or price_fin <= 0:
        return 0.0
    sbc = find_row(cashflow, ['Stock Based Compensation', 'StockBasedCompensation',
                              'Stock Based Compensation Expense'], 0.0)
    if sbc <= 0:
        return 0.0
    dil = (sbc / shares) / price_fin
    return max(0.0, min(0.06, dil))

def revenue_margin_value(rev_ps0, g0, r, g_term, target_margin, exit_pe,
                         cur_margin, dilution=0.0, years=12, plateau=4):
    """Value a pre-profit company off projected revenue.

    Hyper-growth names sustain high growth for years before maturing, so growth
    holds at `g0` for `plateau` years, then fades linearly to `g_term` by year
    `years` (a flat fade-from-year-1 badly undervalues them). The net margin
    ramps from today's `cur_margin` to the sector `target_margin`; positive
    interim earnings are discounted at `r` and year-N earnings capitalized at
    `exit_pe`. Per-share value is deflated for annual `dilution`. Statement-
    currency in, statement-currency out. Returns (value, tv_pct) or (None, None)."""
    if rev_ps0 is None or rev_ps0 <= 0 or r <= g_term:
        return None, None
    rev = rev_ps0
    pv_earnings = 0.0
    for t in range(1, years + 1):
        if t <= plateau:
            g = g0
        else:
            g = g0 - (g0 - g_term) * ((t - plateau) / (years - plateau))
        rev *= (1 + g)
        m = cur_margin + (target_margin - cur_margin) * (t / years)
        eps_t = rev * m
        if eps_t > 0:
            pv_earnings += eps_t / ((1 + r) ** t)
    eps_n = rev * target_margin
    tv_pv = (eps_n * exit_pe) / ((1 + r) ** years)
    total = pv_earnings + tv_pv
    if total <= 0:
        return None, None
    total /= ((1 + dilution) ** years)
    tv_pct = tv_pv / (pv_earnings + tv_pv)
    return total, tv_pct

# ---------- RESIDUAL-INCOME / JUSTIFIED-P/B MODEL (banks & insurers) ----------
# Banks have no meaningful free cash flow (deposits/loans dominate), so growing
# an FCFE/NI stream with a Gordon terminal massively overstates them (JPM showed
# +57%, BARC.L +325%). Instead value them off equity: book value plus the
# present value of *excess* returns over the cost of equity. With ROE=r the bank
# is worth exactly book (P/B=1); a durable ROE>r earns a premium to book.
def residual_income_value(bv0, roe0, r, g_term, payout, years=12,
                          roe_sustainable=0.15):
    """Two-stage residual-income equity value per share (statement currency).

    Excess return ROE−r is earned on a book value that compounds at the
    retention rate. ROE fades linearly from `roe0` to a `roe_sustainable`
    long-run level over `years`; book grows by retained earnings each year. The
    final-year residual income is capitalized as a perpetuity growing at
    `g_term`. Returns (value, terminal_pct) or (None, None) on unusable inputs."""
    if not bv0 or bv0 <= 0 or roe0 is None or r <= g_term:
        return None, None
    roe_term = max(r + 0.01, min(roe0, roe_sustainable))
    retention = max(0.0, min(1.0, 1.0 - (payout or 0.0)))
    bv = bv0
    pv_ri = 0.0
    for t in range(1, years + 1):
        roe_t = roe0 - (roe0 - roe_term) * (t / years)
        pv_ri += ((roe_t - r) * bv) / ((1 + r) ** t)
        bv = bv * (1 + roe_t * retention)
    ri_term = (roe_term - r) * bv
    tv_pv = (ri_term * (1 + g_term) / (r - g_term)) / ((1 + r) ** years) if ri_term > 0 else 0.0
    total = bv0 + pv_ri + tv_pv
    if total <= 0:
        return None, None
    tv_pct = tv_pv / total if total > 0 else None
    return total, tv_pct

def bank_inputs_usable(bv, roe, eps):
    """Guard yfinance's occasional garbage book/ROE for financials. Require a
    positive book value and ROE, and (when EPS is available) that implied
    earnings ROE·BV roughly agree with trailing EPS — else fall back to FCFE."""
    if not bv or bv <= 0 or roe is None or roe <= 0:
        return False
    if eps and eps > 0:
        implied = roe * bv
        if implied <= 0 or implied / eps > 3.0 or implied / eps < 0.33:
            return False
    return True

# ---------- MULTIPLES CROSS-CHECK + BLENDED FAIR VALUE ----------
# Peer-multiple P/E by sector — a sanity anchor for "what is the equity worth at
# a normal earnings multiple?". Used to reconcile DCF extremes (fintech float
# inflates FCFE; data glitches deflate it) against analyst targets via a
# confidence-weighted blend, so no single bad input drives the headline.
MULTIPLE_PE = {
    "Technology": 25.0, "Communication Services": 22.0, "Healthcare": 22.0,
    "Consumer Cyclical": 20.0, "Consumer Defensive": 20.0, "Industrials": 22.0,
    "Energy": 12.0, "Basic Materials": 14.0, "Real Estate": 18.0,
    "Utilities": 16.0, "Financial Services": 13.0,
}
DEFAULT_PE = 20.0

def multiples_fair_value(info, sector, fin_ccy, px_ccy):
    """Peer-multiple equity value per share: normalized EPS × sector P/E,
    converted to the price currency. EPS is reported in the statement currency.
    Returns None when there's no positive earnings (e.g. pre-profit names)."""
    eps_cands = [safe_scalar(info.get(k)) for k in ('forwardEps', 'trailingEps')]
    eps_cands = [e for e in eps_cands if e and e > 0]
    if not eps_cands:
        return None
    norm_eps = float(np.median(eps_cands))
    pe = MULTIPLE_PE.get(sector, DEFAULT_PE)
    return convert_currency(norm_eps * pe, fin_ccy, px_ccy)

BLEND_BASE_W = {"intrinsic": 0.45, "analyst": 0.35, "multiples": 0.20}

def data_confidence(info, intrinsic, multiples, analyst, tv_pct=None,
                    fx_applied=False, pre_profit=False):
    """Score how trustworthy each valuation leg's *inputs* are. Returns
    (overall_0_100, grade, factors) where `factors` are 0..1 reliability weights
    for the intrinsic / analyst / multiples legs of the blend. The headline then
    leans on whichever legs are actually well-supported by the data."""
    # Analyst leg: more covering analysts ⇒ more reliable consensus target.
    n_an = safe_scalar(info.get('numberOfAnalystOpinions'), 0)
    analyst_f = 0.0 if not analyst or analyst <= 0 else min(1.0, 0.25 + n_an / 20.0)

    # Multiples leg: needs positive EPS; trailing/forward agreement adds trust.
    te = safe_scalar(info.get('trailingEps'), 0.0)
    fe = safe_scalar(info.get('forwardEps'), 0.0)
    pos = [e for e in (te, fe) if e and e > 0]
    if not multiples or multiples <= 0 or not pos:
        multiples_f = 0.0
    elif len(pos) == 2:
        multiples_f = 0.5 + 0.5 * (min(pos) / max(pos))  # 1.0 when trailing≈forward
    else:
        multiples_f = 0.5

    # Intrinsic leg: penalize a fragile terminal value, divergence from the
    # multiples/analyst consensus, a coarse pre-profit model, and FX conversion.
    if not intrinsic or intrinsic <= 0:
        intrinsic_f = 0.0
    else:
        intrinsic_f = 1.0
        others = [v for v in (multiples, analyst) if v and v > 0]
        if others:
            ratio = intrinsic / float(np.median(others))
            if ratio > 2.5 or ratio < 0.4:
                intrinsic_f *= 0.5
            elif ratio > 1.75 or ratio < 0.57:
                intrinsic_f *= 0.75
        if tv_pct is not None and tv_pct > 0.75:
            intrinsic_f *= 0.6
        if pre_profit:
            intrinsic_f *= 0.7
        if fx_applied:
            intrinsic_f *= 0.9

    factors = {"intrinsic": intrinsic_f, "analyst": analyst_f, "multiples": multiples_f}
    overall = 100.0 * sum(BLEND_BASE_W[k] * factors[k] for k in factors)
    grade = ("A" if overall >= 80 else "B" if overall >= 65 else
             "C" if overall >= 50 else "D" if overall >= 35 else "E")
    return overall, grade, factors

def blended_fair_value(intrinsic, multiples, analyst, factors=None):
    """Blend the intrinsic value, a peer-multiple value, and the analyst target
    into the headline fair value. Each leg's base importance is scaled by a
    0..1 data-confidence `factor`, so unreliable legs (a fragile/divergent DCF,
    thin analyst coverage, no clean EPS) fade out and can't drive the headline.
    Returns (blend, weights_dict)."""
    vals = {"intrinsic": intrinsic, "analyst": analyst, "multiples": multiples}
    factors = factors or {k: 1.0 for k in BLEND_BASE_W}
    weights = {k: BLEND_BASE_W[k] * factors.get(k, 1.0)
               for k, v in vals.items() if v and v > 0}
    tot = sum(weights.values())
    if tot <= 0:
        return intrinsic, {}
    weights = {k: w / tot for k, w in weights.items()}
    blend = sum(vals[k] * w for k, w in weights.items())
    return blend, weights

# ---------- IMPLIED GROWTH (Reverse DCF) ----------
def implied_growth(current_price, base_fcfe, r, g_term, max_growth=0.40):
    """Solve for the stage-1 growth rate that makes the FCFE DCF equal today's
    price (what the market is implicitly pricing in)."""
    if current_price <= 0 or base_fcfe <= 0:
        return None
    low, high = -0.20, max_growth
    for _ in range(40):
        mid = (low + high) / 2
        val = fcfe_two_stage(base_fcfe, mid, r, g_term)
        if val > current_price:
            high = mid
        else:
            low = mid
    return (low + high) / 2

# ---------- MONTE CARLO DCF ----------
def monte_carlo_dcf(base_fcfe, g_start, r, g_term, growth_std=0.03, r_std=0.01, n_sims=5000):
    values = []
    growths = np.random.normal(g_start, growth_std, n_sims)
    rates = np.random.normal(r, r_std, n_sims)
    for g, rr in zip(growths, rates):
        g = max(-0.10, min(0.50, g))
        rr = max(g_term + 0.01, rr)
        values.append(fcfe_two_stage(base_fcfe, g, rr, g_term))
    return np.percentile(values, [5, 50, 95]), np.std(values)

# ---------- ANALYST RATINGS ----------
def _empty_ratings():
    return {"strong_buy": 0, "buy": 0, "hold": 0, "sell": 0, "strong_sell": 0}

def get_analyst_ratings(ticker):
    """Analyst recommendation counts + price targets.

    yfinance's `recommendations` is a DataFrame with columns
    [period, strongBuy, buy, hold, sell, strongSell], one row per recent month
    ('0m' = current). We read the latest month. Falls back to the legacy
    grade-history format, then to `info`'s recommendationKey, if present."""
    counts = _empty_ratings()
    try:
        rec = ticker.recommendations
        if rec is not None and not rec.empty:
            cols = {'strongBuy': 'strong_buy', 'buy': 'buy', 'hold': 'hold',
                    'sell': 'sell', 'strongSell': 'strong_sell'}
            if set(cols).issubset(rec.columns):
                # Current-month row (period == '0m'), else the first row.
                row = rec[rec['period'] == '0m']
                row = row.iloc[0] if not row.empty else rec.iloc[0]
                for src, dst in cols.items():
                    counts[dst] = int(safe_scalar(row[src]))
            else:
                # Legacy grade-history format.
                grade_col = next((c for c in ['To Grade', 'toGrade', 'Grade', 'Action', 'Rating'] if c in rec.columns), None)
                if grade_col:
                    for grade in rec[grade_col].dropna():
                        g = str(grade).lower()
                        if 'strong buy' in g or 'strong_buy' in g:
                            counts['strong_buy'] += 1
                        elif 'buy' in g:
                            counts['buy'] += 1
                        elif 'hold' in g or 'neutral' in g:
                            counts['hold'] += 1
                        elif 'strong sell' in g or 'strong_sell' in g:
                            counts['strong_sell'] += 1
                        elif 'sell' in g:
                            counts['sell'] += 1
    except Exception:
        pass
    info = {}
    try:
        info = ticker.info
    except Exception:
        pass
    return (counts, info.get('targetMeanPrice'), info.get('targetMedianPrice'),
            info.get('targetHighPrice'), info.get('targetLowPrice'))

# ---------- PEER PERCENTILE RANKING ----------
def get_peer_percentiles(ticker_symbol, sector, current_pe, current_ev_ebitda):
    # Simplified: use predefined sector peers
    sector_peers = {
        "Technology": ["AAPL", "MSFT", "GOOGL", "NVDA", "ADBE", "CRM", "ORCL", "IBM", "CSCO", "INTC"],
        "Financial Services": ["JPM", "BAC", "WFC", "C", "GS", "MS", "V", "MA", "AXP", "BLK"],
        "Healthcare": ["JNJ", "PFE", "MRK", "UNH", "ABBV", "LLY", "AMGN", "GILD", "BMY", "MDT"],
        "Consumer Cyclical": ["AMZN", "TSLA", "HD", "MCD", "NKE", "SBUX", "TGT", "LOW", "BKNG", "EBAY"],
        "Energy": ["XOM", "CVX", "COP", "EOG", "SLB", "PSX", "MPC", "VLO", "OXY", "HES"],
        "Industrials": ["GE", "CAT", "HON", "UPS", "BA", "LMT", "RTX", "UNP", "DE", "MMM"],
        "Real Estate": ["PLD", "AMT", "CCI", "EQIX", "SPG", "O", "DLR", "PSA", "WELL", "AVB"],
        "Basic Materials": ["LIN", "APD", "DOW", "DD", "NEM", "FCX", "SHW", "ECL", "PPG", "ALB"],
        "Communication Services": ["META", "GOOG", "NFLX", "DIS", "TMUS", "VZ", "T", "CMCSA", "CHTR", "ATVI"]
    }
    peers = sector_peers.get(sector, ["AAPL", "MSFT", "GOOGL"])
    pe_vals = []
    ev_vals = []
    for p in peers:
        if p == ticker_symbol.upper():
            continue
        try:
            pinfo = yf.Ticker(p).info
            pe = pinfo.get('trailingPE')
            ev = pinfo.get('enterpriseToEbitda')
            if pe and pe > 0:
                pe_vals.append(pe)
            if ev and ev > 0:
                ev_vals.append(ev)
        except Exception:
            pass
    if not pe_vals:
        return 50, 50
    pe_percentile = sum(1 for x in pe_vals if x < current_pe) / len(pe_vals) * 100 if current_pe else 50
    ev_percentile = sum(1 for x in ev_vals if x < current_ev_ebitda) / len(ev_vals) * 100 if current_ev_ebitda else 50
    return pe_percentile, ev_percentile

# ---------- TECHNICAL INDICATORS ----------
def get_technicals(ticker, period="6mo"):
    hist = ticker.history(period=period)
    if hist.empty:
        return {}
    close = hist['Close'].apply(safe_scalar)
    sma_50 = safe_scalar(close.rolling(50).mean().iloc[-1] if len(close) >= 50 else close.mean())
    sma_200 = safe_scalar(close.rolling(200).mean().iloc[-1] if len(close) >= 200 else sma_50)
    delta = close.diff()
    gain = delta.where(delta > 0, 0).rolling(window=14).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
    rs = gain / loss
    rsi = safe_scalar(100 - (100 / (1 + rs)).iloc[-1] if not rs.empty else 50)
    exp1 = close.ewm(span=12, adjust=False).mean()
    exp2 = close.ewm(span=26, adjust=False).mean()
    macd = exp1 - exp2
    signal = macd.ewm(span=9, adjust=False).mean()
    macd_hist = safe_scalar((macd - signal).iloc[-1] if not macd.empty else 0)
    try:
        spy_df = yf.download("SPY", period=period, progress=False, auto_adjust=True)
        spy = spy_df['Close']
        # Newer yfinance returns multi-index columns -> 'Close' is a DataFrame.
        if isinstance(spy, pd.DataFrame):
            spy = spy.iloc[:, 0]
    except Exception:
        spy = pd.Series(dtype=float)
    if not spy.empty:
        stock_ret = safe_scalar(close.pct_change().iloc[-20:].mean() * 252)
        spy_ret = safe_scalar(spy.pct_change().iloc[-20:].mean() * 252)
        rel_strength = stock_ret / spy_ret if spy_ret != 0 else 1.0
    else:
        rel_strength = 1.0
    return {"sma_50": sma_50, "sma_200": sma_200, "rsi": rsi, "macd_hist": macd_hist, "rel_strength": rel_strength}

# ---------- EXTRA METRICS ----------
def get_extra_metrics(ticker, info, financials, cashflow, balancesheet, shares, current_price):
    metrics = {}
    try:
        cfo = safe_scalar(cashflow.loc['Operating Cash Flow'].iloc[0]) if 'Operating Cash Flow' in cashflow.index else 0
        net_income = safe_scalar(financials.loc['Net Income'].iloc[0]) if 'Net Income' in financials.index else 0
        if net_income != 0:
            metrics['earnings_quality'] = cfo / net_income if cfo != 0 else None
        total_equity = safe_scalar(balancesheet.loc['Total Equity Gross Minority Interest'].iloc[0]) if 'Total Equity Gross Minority Interest' in balancesheet.index else 1
        if total_equity != 0 and net_income != 0:
            roe = net_income / total_equity
            dividends = safe_scalar(cashflow.loc['Dividends Paid'].iloc[0]) if 'Dividends Paid' in cashflow.index else 0
            payout = dividends / net_income if net_income != 0 else 0
            sustainable = roe * (1 - payout)
            metrics['sustainable_growth'] = min(0.20, sustainable) if sustainable > 0 else None
        shares_begin = info.get('sharesOutstanding', shares)
        buyback_yield = ((shares_begin - shares) / shares_begin) if shares_begin else 0
        metrics['buyback_yield'] = max(0, buyback_yield) * 100
        div_yield = info.get('dividendYield', 0) * 100 if info.get('dividendYield') else 0
        metrics['dividend_yield'] = div_yield
        metrics['total_shareholder_yield'] = div_yield + metrics['buyback_yield']
        fcf_abs = cfo + (safe_scalar(cashflow.loc['Capital Expenditure'].iloc[0]) if 'Capital Expenditure' in cashflow.index else 0)
        metrics['fcf_yield'] = (fcf_abs / (current_price * shares)) * 100 if current_price and shares else None
        return metrics
    except Exception:
        return {}

# ---------- CONFIDENCE SCORE ----------
def compute_confidence(mos, analyst_counts, technical, extra):
    cfg = load_config()
    w = cfg['confidence_weights']
    mos_norm = max(0, min(1, (mos + 50) / 100))
    total = sum(analyst_counts.values())
    if total > 0:
        analyst_score = (analyst_counts['strong_buy']*1.0 + analyst_counts['buy']*0.7 +
                         analyst_counts['hold']*0.3 - analyst_counts['sell']*0.5 - analyst_counts['strong_sell']*1.0) / total
        analyst_norm = max(0, min(1, (analyst_score + 1) / 2))
    else:
        analyst_norm = 0.5
    tech_score = 0.5
    if technical:
        rsi = technical.get('rsi', 50)
        if 30 <= rsi <= 70:
            tech_score += 0.2
        elif rsi < 30:
            tech_score += 0.1
        if technical.get('macd_hist', 0) > 0:
            tech_score += 0.15
        if technical.get('rel_strength', 1) > 1:
            tech_score += 0.15
        tech_norm = min(1, tech_score)
    else:
        tech_norm = 0.5
    qual_score = 0.5
    if extra:
        eq = extra.get('earnings_quality')
        if eq and eq > 0.8:
            qual_score += 0.2
        sg = extra.get('sustainable_growth')
        if sg and sg > 0.05:
            qual_score += 0.15
        if extra.get('total_shareholder_yield', 0) > 3:
            qual_score += 0.15
        qual_norm = min(1, qual_score)
    else:
        qual_norm = 0.5
    confidence = w['mos']*mos_norm + w['analyst']*analyst_norm + w['technical']*tech_norm + w['quality']*qual_norm
    return confidence * 100

# ---------- MAIN ANALYSIS (v12) ----------
def analyze_ticker(ticker_symbol: str, use_cache: bool = True, base_mode: str = None):
    # base_mode None => use the configured default, but allow a sector-aware
    # override below (financials value off net income, not operating cash flow).
    explicit_base = base_mode is not None
    cache_key = f"analysis_{ticker_symbol.upper()}_{base_mode or 'default'}"
    if use_cache:
        cached = cache_get(cache_key, expiry_hours=6)
        if cached:
            console.print(f"[dim]✓ Loaded {ticker_symbol.upper()} from cache (<6h old)[/dim]")
            return cached
    try:
        ticker = yf.Ticker(ticker_symbol)
        info = ticker.info
        financials = ticker.financials
        cashflow = ticker.cashflow
        balancesheet = ticker.balance_sheet

        current_price = info.get('currentPrice', info.get('previousClose', 0))
        shares = info.get('sharesOutstanding', 0)
        if shares == 0 or current_price == 0:
            return None

        # ----- Fundamental data -----
        # Normalize FCF over several years so a single capex spike doesn't wreck
        # the DCF (e.g. AMZN's data-center buildout collapsed one-year FCF).
        _, n_years = normalized_fcf_per_share(cashflow, shares, years=3)
        if n_years >= 2:
            console.print(f"[dim]✓ FCF base: {n_years}-yr average (smoothed)[/dim]")

        # ----- Growth seed & discount rate (cost of equity) -----
        cfg = load_config()
        growth = estimate_growth(ticker, financials, cashflow, info, shares)
        g_start = analyst_growth_rate(ticker, growth)
        growth_cap = cfg.get('max_stage1_growth', 0.20)
        g_start = max(cfg['min_growth'], min(growth_cap, g_start))
        rf = get_risk_free_rate()
        erp = cfg.get('equity_risk_premium', get_market_risk_premium())
        beta = max(0.4, info.get('beta', 1.0) or 1.0)
        crp = country_risk_premium(info.get('country'))
        discount = cost_of_equity(beta, rf, erp, crp)
        g_term = get_longrun_growth()
        if g_term >= discount:
            g_term = discount - 0.02

        # ----- Resolve the base cash-flow mode (sector-aware) -----
        # Banks/insurers don't have a meaningful "operating cash flow" (it's
        # dominated by loan/deposit flows), so unless the user explicitly chose
        # a base, financials value off net income instead of the OCF default.
        sector = info.get('sector', 'Technology')
        is_bank = sector == 'Financial Services'
        if not explicit_base:
            base_mode = load_config().get('default_base', 'ocf')
            if is_bank and base_mode == 'ocf':
                base_mode = 'ni'

        # ----- Currency normalization -----
        # Statements are in `financialCurrency`; the quoted price is in
        # `currency` (often a different currency for ADRs, or pence for `.L`).
        # We value in the statement currency, then convert per-share values into
        # the price currency so the upside/MOS comparison is apples-to-apples.
        fin_ccy = info.get('financialCurrency') or info.get('currency')
        px_ccy = info.get('currency') or fin_ccy
        fx_applied = bool(fin_ccy and px_ccy and fin_ccy != px_ccy)

        # ----- 2-stage FCFE valuations (base, bear, bull) -----
        base_fcfe = normalized_cash_earnings_ps(cashflow, financials, shares, mode=base_mode)
        # Net cash per share (cash - debt), a non-operating asset the FCFE stream
        # doesn't capture. Skipped for banks, where debt/deposits are operational.
        net_cash_ps = 0.0
        if cfg.get('add_net_cash', True) and not is_bank:
            total_cash = safe_scalar(info.get('totalCash'), 0.0)
            total_debt = safe_scalar(info.get('totalDebt'), 0.0)
            if shares:
                net_cash_ps = (total_cash - total_debt) / shares

        pre_profit = base_fcfe <= 0
        nc_px = convert_currency(net_cash_ps, fin_ccy, px_ccy)
        price_fin = convert_currency(current_price, px_ccy, fin_ccy)
        bank_bv = safe_scalar(info.get('bookValue'), 0.0)
        bank_roe = info.get('returnOnEquity')
        bank_payout = info.get('payoutRatio')
        use_bank = is_bank and bank_inputs_usable(bank_bv, bank_roe, info.get('trailingEps'))
        if use_bank:
            # Banks: value off book + excess returns (residual income), not FCFE.
            method = "residual_income"
            pre_profit = False

            def _bank_val(roe, sustain):
                v, tvp = residual_income_value(bank_bv, roe, discount, g_term,
                                               bank_payout, roe_sustainable=sustain)
                if v is None:
                    return None, None
                return convert_currency(v, fin_ccy, px_ccy), tvp
            val_base, tv_pct = _bank_val(bank_roe, 0.15)
            val_bear, _ = _bank_val(bank_roe * 0.85, 0.12)
            val_bull, _ = _bank_val(bank_roe * 1.10, 0.17)
            mos = ((val_base - current_price) / current_price) * 100 if (val_base and current_price > 0) else None
            implied_g = None
        elif pre_profit:
            # FCFE is undefined for pre-profit names (negative base => $0). Value
            # the *future* profitable business off projected revenue instead.
            method = "rev_margin"
            rev_ps0 = find_row(financials, ['Total Revenue', 'TotalRevenue',
                                            'Operating Revenue'], 0.0) / shares
            g0 = revenue_growth_seed(info, financials, cap_hi=0.50)
            tgt_margin, exit_pe = sector_margin_exit(sector)
            cur_margin = max(-0.5, min(tgt_margin, safe_scalar(info.get('profitMargins'), 0.0)))
            dil = annual_dilution(cashflow, financials, shares, price_fin)

            def _rev_val(g, m, pe):
                v, tvp = revenue_margin_value(rev_ps0, g, discount, g_term, m, pe,
                                              cur_margin, dilution=dil)
                if v is None:
                    return None, None
                return convert_currency(v, fin_ccy, px_ccy) + nc_px, tvp
            val_base, tv_pct = _rev_val(g0, tgt_margin, exit_pe)
            val_bear, _ = _rev_val(g0 * 0.7, tgt_margin * 0.7, exit_pe * 0.8)
            val_bull, _ = _rev_val(min(0.60, g0 * 1.2), tgt_margin * 1.2, exit_pe * 1.15)
            mos = ((val_base - current_price) / current_price) * 100 if (val_base and current_price > 0) else None
            implied_g = None
        else:
            method = "fcfe"
            def _val(g, r):
                pv, tv_pv = fcfe_two_stage_parts(base_fcfe, g, r, g_term)
                op = pv + tv_pv
                op_px = convert_currency(op, fin_ccy, px_ccy)
                return op_px + nc_px, (tv_pv / op if op > 0 else None)
            val_base, tv_pct = _val(g_start, discount)
            val_bear, _ = _val(g_start * 0.6, discount + 0.015)
            val_bull, _ = _val(g_start * 1.25, max(g_term + 0.02, discount - 0.01))
            # Upside/downside vs. price (bounded, intuitive).
            mos = ((val_base - current_price) / current_price) * 100 if current_price > 0 else 0
            # ----- Implied growth (reverse DCF) -----
            # Compare in statement currency: convert the price back, net of cash.
            implied_g = implied_growth(max(0.0, price_fin - net_cash_ps), base_fcfe, discount, g_term)

        # ----- Monte Carlo (if enabled) -----
        mc_lower = mc_median = mc_upper = None
        if cfg['use_monte_carlo'] and not pre_profit and not use_bank:
            percentiles, _ = monte_carlo_dcf(base_fcfe, g_start, discount, g_term, n_sims=cfg['num_simulations'])
            nc_px = convert_currency(net_cash_ps, fin_ccy, px_ccy)
            mc_lower, mc_median, mc_upper = [
                convert_currency(float(p), fin_ccy, px_ccy) + nc_px for p in percentiles
            ]

        # ----- Analyst & Short, Options -----
        analyst_counts, t_mean, t_median, t_high, t_low = get_analyst_ratings(ticker)
        shares_short = info.get('sharesShort', 0)
        float_shares = info.get('floatShares', shares)
        short_float = (shares_short / float_shares * 100) if float_shares and float_shares > 0 else None
        short_ratio = info.get('shortRatio', None)

        # Options put/call ratio
        try:
            expirations = ticker.options
            put_call_vol = None
            if expirations:
                chain = ticker.option_chain(expirations[0])
                calls_vol = safe_scalar(chain.calls['volume'].sum())
                puts_vol = safe_scalar(chain.puts['volume'].sum())
                put_call_vol = puts_vol / calls_vol if calls_vol != 0 else None
        except Exception:
            put_call_vol = None

        # ----- Quality & Technicals -----
        piotroski = None  # simplified, left as previous
        technical = get_technicals(ticker, period="1y")
        extra = get_extra_metrics(ticker, info, financials, cashflow, balancesheet, shares, current_price)
        confidence = compute_confidence(mos if mos is not None else 0.0, analyst_counts, technical, extra)

        # ----- Multiples cross-check + blended (headline) fair value -----
        # Reconcile the intrinsic value with a peer-multiple value and the
        # analyst target so float-inflated DCFs (MELI), data glitches (Nintendo)
        # or fragile terminal values don't drive the headline on their own.
        multiples_val = multiples_fair_value(info, sector, fin_ccy, px_ccy)
        data_conf, data_grade, conf_factors = data_confidence(
            info, val_base, multiples_val, t_mean, tv_pct=tv_pct,
            fx_applied=fx_applied, pre_profit=pre_profit)
        blended_val, blend_weights = blended_fair_value(val_base, multiples_val, t_mean, factors=conf_factors)
        blended_mos = ((blended_val - current_price) / current_price) * 100 if (blended_val and current_price > 0) else None

        # ----- Peer percentiles -----
        pe = info.get('trailingPE', 0)
        ev_ebitda = info.get('enterpriseToEbitda', 0)
        pe_percentile, ev_percentile = get_peer_percentiles(ticker_symbol, sector, pe, ev_ebitda)

        # ----- Risk Flags -----
        flags = []
        if extra.get('earnings_quality') and extra['earnings_quality'] < 0.8:
            flags.append("⚠️ Low earnings quality (CFO < NI)")
        if extra.get('fcf_yield') and extra['fcf_yield'] < 2:
            flags.append("⚠️ Low FCF yield (<2%)")
        if short_float and short_float > 20:
            flags.append("📉 High short interest >20%")
        if pre_profit:
            flags.append("🌱 Pre-profit: valued via revenue→margin model (no current FCF)")
        if use_bank:
            flags.append("🏦 Bank: valued via residual income (book value + excess ROE)")
        if val_base is not None and val_base < current_price * 0.7:
            flags.append("🔴 Deeply overvalued (MOS < -30%)")
        if put_call_vol and put_call_vol > 1.2:
            flags.append("🐻 Elevated put/call ratio")
        if implied_g and implied_g > 0.15:
            flags.append(f"📈 Market expects {implied_g:.0%} growth (unrealistic)")
        if fx_applied:
            flags.append(f"💱 FX-normalized: statements {fin_ccy} → price {px_ccy}")
        if tv_pct is not None and tv_pct > 0.75:
            flags.append(f"⏳ {tv_pct:.0%} of value is terminal (fragile)")
        if blended_val and val_base and val_base > 0 and (val_base / blended_val > 1.5 or val_base / blended_val < 0.67):
            flags.append("⚖️ Intrinsic diverges from peers/analysts; headline is a blend")

        result = {
            "ticker": ticker_symbol.upper(),
            "name": info.get('longName', ticker_symbol),
            "price": current_price,
            "dcf": {"bear": val_bear, "base": val_base, "bull": val_bull},
            "multiples_value": multiples_val,
            "blended_value": blended_val,
            "blended_weights": blend_weights,
            "blended_mos": blended_mos,
            "data_confidence": data_conf,
            "data_grade": data_grade,
            "conf_factors": conf_factors,
            "mos": mos,
            "confidence": confidence,
            "growth_used": g_start,
            "wacc_used": discount,
            "discount_rate": discount,
            "country_risk_premium": crp,
            "terminal_growth": g_term,
            "base_fcfe": base_fcfe,
            "base_mode": base_mode,
            "method": method,
            "book_value": bank_bv if use_bank else None,
            "roe": bank_roe if use_bank else None,
            "currency": px_ccy,
            "financial_currency": fin_ccy,
            "fx_applied": fx_applied,
            "net_cash_ps": net_cash_ps,
            "pre_profit": pre_profit,
            "tv_pct": tv_pct,
            "implied_growth": implied_g,
            "mc": {"lower": mc_lower, "median": mc_median, "upper": mc_upper} if mc_lower else None,
            "analyst": {"ratings": dict(analyst_counts), "target_mean": t_mean, "target_median": t_median, "target_high": t_high, "target_low": t_low},
            "short": {"float_pct": short_float, "days_to_cover": short_ratio, "shares": shares_short},
            "options": {"put_call_vol": put_call_vol},
            "quality": {"piotroski": piotroski},
            "technical": technical,
            "extra": extra,
            "peer_percentiles": {"pe": pe_percentile, "ev_ebitda": ev_percentile},
            "risk_flags": flags
        }
        result = to_json_safe(result)
        if use_cache:
            cache_set(cache_key, result)
        return result
    except Exception as e:
        console.print(f"[red]Analysis error for {ticker_symbol}: {e}[/red]")
        return None

# ---------- RENDER DASHBOARD (v12) ----------
def render_dashboard(res):
    console.clear()
    console.print(Panel.fit(f"[bold gold1]🧠 AEGIS OMNISCIENT v12.0 • {res['name']} ({res['ticker']}) • 1Y+ Valuation[/bold gold1]"))

    # Valuation table
    cur = res.get('currency') or 'USD'
    def _money(v):
        return f"{cur} {v:.2f}" if v is not None else "[dim]n/a[/dim]"
    val_table = Table(title="💎 Intrinsic Value", box=None)
    val_table.add_column("Scenario", style="cyan")
    val_table.add_column("Target")
    val_table.add_row("Bear (Stress)", _money(res['dcf']['bear']))
    val_table.add_row("Base (Intrinsic)", _money(res['dcf']['base']))
    val_table.add_row("Bull (Optimistic)", f"[bold green]{_money(res['dcf']['bull'])}[/bold green]")
    if res.get('multiples_value') is not None:
        val_table.add_row("Peer Multiple", _money(res['multiples_value']))
    if res['analyst'].get('target_mean'):
        val_table.add_row("Analyst Target", _money(res['analyst']['target_mean']))
    if res['mc']:
        val_table.add_row("Monte Carlo (5%-95%)", f"{_money(res['mc']['lower'])} → {_money(res['mc']['median'])} → {_money(res['mc']['upper'])}")
    blended = res.get('blended_value')
    if blended is not None:
        val_table.add_row("[bold]Blended Fair Value[/bold]", f"[bold gold1]{_money(blended)}[/bold gold1]")
    if res.get('data_confidence') is not None:
        dc = res['data_confidence']
        dc_color = "green" if dc >= 65 else "yellow" if dc >= 50 else "red"
        val_table.add_row("Data Confidence", f"[{dc_color}]{dc:.0f}/100 ({res.get('data_grade', '-')})[/{dc_color}]")
    bmos = res.get('blended_mos')
    if bmos is not None:
        val_table.add_row("Upside vs Price", f"{bmos:.1f}%" + (" 🟢" if bmos > 20 else " 🟡" if bmos > 0 else " 🔴"))
    elif res.get('mos') is not None:
        mos = res['mos']
        val_table.add_row("Upside vs Price", f"{mos:.1f}%" + (" 🟢" if mos > 20 else " 🟡" if mos > 0 else " 🔴"))
    else:
        val_table.add_row("Upside vs Price", "[dim]n/a (pre-profit)[/dim]")
    val_table.add_row("Confidence", f"{safe_scalar(res['confidence']):.0f}/100")

    # Analyst & Market
    anal_table = Table(title="📊 Analyst & Market", box=None)
    anal_table.add_column("Metric", style="cyan")
    anal_table.add_column("Value")
    r = res['analyst']['ratings']
    total_ratings = sum(r.values())
    rating_str = f"SB:{r['strong_buy']} B:{r['buy']} H:{r['hold']} S:{r['sell']} SS:{r['strong_sell']}"
    if total_ratings:
        rating_str += f"  (n={total_ratings})"
    anal_table.add_row("Ratings", rating_str)
    tgt = res['analyst']['target_mean']
    price = res.get('price', 0)
    if tgt:
        line = f"${safe_scalar(tgt):.2f}"
        if price and price > 0:
            up = (safe_scalar(tgt) - price) / price * 100
            line += f"  ({up:+.1f}% vs price)"
        anal_table.add_row("Analyst Target", line)
    if res['short']['float_pct']:
        anal_table.add_row("Short Float", f"{safe_scalar(res['short']['float_pct']):.1f}%")
    if res['options']['put_call_vol']:
        anal_table.add_row("Put/Call", f"{safe_scalar(res['options']['put_call_vol']):.2f}")
    anal_table.add_row("P/E vs Peers", f"{res['peer_percentiles']['pe']:.0f}th percentile")
    anal_table.add_row("EV/EBITDA vs Peers", f"{res['peer_percentiles']['ev_ebitda']:.0f}th percentile")

    # Quality & Efficiency
    qual_table = Table(title="🧪 Quality", box=None)
    qual_table.add_column("Metric", style="cyan")
    qual_table.add_column("Value")
    eq = res['extra'].get('earnings_quality')
    if eq:
        qual_table.add_row("Earnings Quality", f"{eq:.2f}")
    sg = res['extra'].get('sustainable_growth')
    if sg:
        qual_table.add_row("Sustainable Growth", f"{sg:.1%}")
    if res['extra'].get('total_shareholder_yield'):
        qual_table.add_row("Total Yield", f"{res['extra']['total_shareholder_yield']:.1f}%")
    if res['extra'].get('fcf_yield'):
        qual_table.add_row("FCF Yield", f"{res['extra']['fcf_yield']:.1f}%")
    if res['growth_used']:
        qual_table.add_row("Assumed Growth", f"{res['growth_used']:.1%}")
    if res['implied_growth']:
        qual_table.add_row("Market Implied Growth", f"{res['implied_growth']:.1%}")

    # Technicals
    tech_table = Table(title="📈 Technicals", box=None)
    tech_table.add_column("Indicator", style="cyan")
    tech_table.add_column("Value")
    tech = res['technical']
    if tech:
        rsi_val = safe_scalar(tech.get('rsi', 50))
        rsi_str = f"{rsi_val:.1f}" + (" (Oversold)" if rsi_val < 30 else " (Overbought)" if rsi_val > 70 else "")
        tech_table.add_row("RSI", rsi_str)
        tech_table.add_row("vs SMA50", f"{safe_scalar(tech.get('sma_50', 0)):.2f}")
        tech_table.add_row("MACD Hist", f"{safe_scalar(tech.get('macd_hist', 0)):.4f}")
        tech_table.add_row("Rel Strength", f"{safe_scalar(tech.get('rel_strength', 1)):.2f}")

    panels = [Panel(val_table), Panel(anal_table), Panel(qual_table), Panel(tech_table)]
    try:
        console.print(Columns(panels, width=120))
    except Exception:
        for p in panels:
            console.print(p)

    if res['risk_flags']:
        console.print(Panel("\n".join(res['risk_flags']), title="⚠️ Risk Flags", border_style="red"))

    # Recommendation
    conf = safe_scalar(res['confidence'])
    if conf >= 80:
        rec = "[bold green]STRONG BUY[/bold green] 🟢"
    elif conf >= 65:
        rec = "[bold cyan]BUY[/bold cyan] 📈"
    elif conf >= 40:
        rec = "[bold yellow]HOLD[/bold yellow] ⚖️"
    elif conf >= 20:
        rec = "[bold orange1]SELL[/bold orange1] 📉"
    else:
        rec = "[bold red]STRONG SELL[/bold red] 🔴"
    console.print(Panel(rec, title="Aegis Conviction", border_style="magenta"))
    base_lbl = {'ni': 'net income', 'ocf': 'operating cash flow', 'fcf': '3yr-avg FCF', 'auto': 'normalized'}.get(res.get('base_mode', 'auto'), 'normalized')
    crp = res.get('country_risk_premium', 0.0)
    crp_lbl = f" (incl. {crp:.1%} country risk)" if crp and crp > 0 else ""
    fin_ccy = res.get('financial_currency')
    fx_lbl = f" | Statements: {fin_ccy}→{cur}" if res.get('fx_applied') else ""
    nc = res.get('net_cash_ps')
    nc_lbl = f" | Net cash: {cur} {nc:+.2f}/sh" if nc else ""
    fv = res.get('blended_value') if res.get('blended_value') is not None else res['dcf']['base']
    fv_lbl = f"{cur} {fv:.2f} (blended)" if res.get('blended_value') is not None else (f"{cur} {fv:.2f}" if fv is not None else "n/a")
    if res.get('method') == 'rev_margin':
        base_part = "Base: revenue→margin (pre-profit)"
    elif res.get('method') == 'residual_income':
        bv = res.get('book_value')
        base_part = f"Base: residual income (book {fin_ccy or cur} {bv:.2f}/sh, ROE {res.get('roe', 0):.1%})" if bv else "Base: residual income (book + excess ROE)"
    else:
        base_part = f"Base: {base_lbl} ({fin_ccy or cur} {res.get('base_fcfe', 0):.2f}/sh)"
    console.print(f"[dim]Discount rate (cost of equity): {res.get('discount_rate', res['wacc_used']):.1%}{crp_lbl} | Terminal g: {res.get('terminal_growth', 0.025):.1%} | {base_part}{nc_lbl}{fx_lbl} | Fair value: {fv_lbl}[/dim]")

# ---------- WATCHLIST HELPERS ----------
def load_watchlist():
    if os.path.exists(WATCHLIST_FILE):
        try:
            with open(WATCHLIST_FILE, "r") as f:
                data = json.load(f)
            if isinstance(data, list):
                return [str(t).upper() for t in data]
        except Exception:
            pass
    return []

def save_watchlist(tickers):
    unique = sorted({str(t).upper() for t in tickers})
    with open(WATCHLIST_FILE, "w") as f:
        json.dump(unique, f, indent=2)
    return unique

# ---------- EXPORT ----------
def export_result(res, fmt):
    """Export a single analysis result. Returns the path written, or None."""
    fmt = fmt.lower()
    ticker = res.get("ticker", "UNKNOWN")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.join(REPORT_DIR, f"{ticker}_{stamp}")
    safe = to_json_safe(res)

    def _flat_rows(r):
        return [
            ("Ticker", r.get("ticker")),
            ("Name", r.get("name")),
            ("Price", r.get("price")),
            ("DCF Bear", r.get("dcf", {}).get("bear")),
            ("DCF Base", r.get("dcf", {}).get("base")),
            ("DCF Bull", r.get("dcf", {}).get("bull")),
            ("Upside vs Price %", r.get("mos")),
            ("Confidence", r.get("confidence")),
            ("Growth Used", r.get("growth_used")),
            ("Discount Rate (CoE)", r.get("discount_rate", r.get("wacc_used"))),
            ("Implied Growth", r.get("implied_growth")),
        ]

    if fmt == "json":
        path = base + ".json"
        with open(path, "w") as f:
            json.dump(safe, f, indent=2)
        return path

    if fmt == "csv":
        import csv
        path = base + ".csv"
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Metric", "Value"])
            for k, v in _flat_rows(safe):
                writer.writerow([k, v])
        return path

    if fmt in ("excel", "xlsx"):
        if not HAS_OPENPYXL:
            console.print("[yellow]openpyxl not installed; skipping Excel export.[/yellow]")
            return None
        path = base + ".xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Valuation"
        ws.append(["Metric", "Value"])
        for k, v in _flat_rows(safe):
            ws.append([k, v])
        wb.save(path)
        return path

    if fmt == "pdf":
        if not HAS_FPDF:
            console.print("[yellow]fpdf2 not installed; skipping PDF export.[/yellow]")
            return None
        path = base + ".pdf"
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, f"Aegis Valuation - {ticker}", ln=True)
        pdf.set_font("Helvetica", size=11)
        for k, v in _flat_rows(safe):
            pdf.cell(0, 8, f"{k}: {v}", ln=True)
        if safe.get("risk_flags"):
            pdf.ln(4)
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, "Risk Flags", ln=True)
            pdf.set_font("Helvetica", size=11)
            for flag in safe["risk_flags"]:
                # fpdf core fonts are latin-1 only; drop unsupported glyphs (emoji).
                clean = flag.encode("latin-1", "ignore").decode("latin-1").strip()
                pdf.cell(0, 8, clean or "- flag", ln=True)
        pdf.output(path)
        return path

    console.print(f"[red]Unknown export format: {fmt}[/red]")
    return None

# ---------- CLI HELPERS (plain functions, callable internally) ----------
def run_single(ticker, export=None, monte_carlo=False, no_cache=False, base_mode=None):
    """Analyze one ticker and render its dashboard. Returns True on success."""
    if not ticker:
        ticker = Prompt.ask("Ticker", default=load_config()["default_ticker"])
    if monte_carlo:
        cfg = load_config()
        cfg["use_monte_carlo"] = True
        save_config(cfg)
    with Progress(SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
                  transient=True) as progress:
        progress.add_task(f"Analyzing {ticker.upper()}...", total=None)
        res = analyze_ticker(ticker, use_cache=not no_cache, base_mode=base_mode)
    if not res:
        console.print(f"[red]Could not analyze {ticker.upper()}. Check the symbol and try again.[/red]")
        return False
    render_dashboard(res)
    if export:
        path = export_result(res, export)
        if path:
            console.print(f"[green]✓ Exported to {path}[/green]")
    return True

def run_compare(tickers, no_cache=False):
    """Analyze multiple tickers and render a comparison table. Returns True on success."""
    results = []
    with Progress(SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
                  transient=True) as progress:
        task = progress.add_task("Analyzing...", total=None)
        for t in tickers:
            progress.update(task, description=f"Analyzing {t.upper()}...")
            res = analyze_ticker(t, use_cache=not no_cache)
            if res:
                results.append(res)
    if not results:
        console.print("[red]No tickers could be analyzed.[/red]")
        return False
    table = Table(title="📊 Comparison")
    table.add_column("Ticker", style="cyan")
    table.add_column("Price")
    table.add_column("Base DCF")
    table.add_column("Upside %")
    table.add_column("Confidence")
    table.add_column("Growth")
    for r in sorted(results, key=lambda x: safe_scalar(x.get("confidence")), reverse=True):
        cur = r.get("currency") or "USD"
        base = r['dcf'].get('base')
        base_str = f"{cur} {base:.2f}" if base is not None else "n/a"
        mos = r.get("mos")
        if mos is None:
            mos_str = "n/a 🌱"
        else:
            mos_str = f"{mos:.1f}%" + (" 🟢" if mos > 20 else " 🟡" if mos > 0 else " 🔴")
        table.add_row(
            r["ticker"],
            f"{cur} {safe_scalar(r.get('price')):.2f}",
            base_str,
            mos_str,
            f"{safe_scalar(r.get('confidence')):.0f}/100",
            f"{safe_scalar(r.get('growth_used')):.1%}",
        )
    console.print(table)
    return True

# ---------- VALIDATION HARNESS ----------
# A diverse basket spanning the archetypes the engine must handle: megacap,
# banks (US/EM/UK), semis, autos, staples/healthcare, pre-profit growth, EM
# high-growth, and foreign listings. Used as a regression check.
VALIDATION_BASKET = [
    "AMZN", "JPM", "NU", "BARC.L", "TSM", "ASML", "RACE", "PEP", "JNJ",
    "OUST", "RKLB", "MELI", "7974.T",
]

def _sanity(res):
    """Heuristic sanity verdict for one analysis result. Returns (label, ok)."""
    if res is None:
        return "no-data", False
    # The headline is the confidence-weighted blend; fall back to intrinsic.
    fv = res.get("blended_value") or res["dcf"].get("base")
    pre = res.get("pre_profit")
    if fv is None and pre:
        return "pre-profit (no rev)", True  # revenue model couldn't seed; tolerated
    if not fv or fv <= 0:
        return "FV<=0", False
    # Compare to analyst target (preferred) or price, in the same currency.
    ref = res["analyst"].get("target_mean") or res.get("price")
    if not ref or ref <= 0:
        return "no reference", True
    ratio = fv / ref
    # Pre-profit names are a deliberately conservative *floor*: a value well
    # below the market's optionality-driven price is expected, so only an
    # implausibly high result is a failure. Profitable names use a tight band.
    lo = 0.10 if pre else 0.40
    if lo <= ratio <= 2.5:
        tag = "floor" if (pre and ratio < 0.4) else "ok"
        return f"{tag} ({ratio:.2f}x)", True
    return f"OUTLIER ({ratio:.2f}x)", False

def run_validate(tickers, no_cache=True):
    """Run the basket and print FV vs price vs analyst target with a sanity
    verdict per name — the regression harness for valuation changes."""
    table = Table(title="🧪 Valuation Validation Harness")
    table.add_column("Ticker", style="cyan")
    table.add_column("Method")
    table.add_column("Ccy")
    table.add_column("Fair Value", justify="right")
    table.add_column("Price", justify="right")
    table.add_column("Analyst Tgt", justify="right")
    table.add_column("Upside", justify="right")
    table.add_column("TV%", justify="right")
    table.add_column("Data", justify="right")
    table.add_column("Verdict")
    n_ok = 0
    n_total = 0
    with Progress(SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
                  transient=True) as progress:
        task = progress.add_task("Validating...", total=None)
        for t in tickers:
            progress.update(task, description=f"Analyzing {t.upper()}...")
            res = analyze_ticker(t, use_cache=not no_cache)
            n_total += 1
            verdict, ok = _sanity(res)
            n_ok += 1 if ok else 0
            if res is None:
                table.add_row(t.upper(), "-", "-", "-", "-", "-", "-", "-", "-", f"[red]{verdict}[/red]")
                continue
            cur = res.get("currency") or "USD"
            method = {
                "rev_margin": "Rev→Margin",
                "residual_income": "Residual",
            }.get(res.get("method"), "FCFE")
            fv = res.get("blended_value") or res["dcf"].get("base")
            fv_s = f"{fv:.2f}" if fv is not None else "n/a"
            tgt = res["analyst"].get("target_mean")
            tgt_s = f"{tgt:.2f}" if tgt else "-"
            mos = res.get("blended_mos") if res.get("blended_mos") is not None else res.get("mos")
            mos_s = f"{mos:+.0f}%" if mos is not None else "n/a"
            tv = res.get("tv_pct")
            tv_s = f"{tv:.0%}" if tv is not None else "-"
            dc = res.get("data_confidence")
            dc_s = f"{dc:.0f} ({res.get('data_grade', '-')})" if dc is not None else "-"
            style = "green" if ok else "red"
            table.add_row(
                res["ticker"], method, cur, fv_s, f"{safe_scalar(res.get('price')):.2f}",
                tgt_s, mos_s, tv_s, dc_s, f"[{style}]{verdict}[/{style}]",
            )
    console.print(table)

    console.print(f"\n[bold]{n_ok}/{n_total} sane[/bold] "
                  f"({'all clear' if n_ok == n_total else 'see OUTLIER/FV<=0 rows'}).")
    return n_ok == n_total

# ---------- CLI COMMANDS ----------
@app.callback(invoke_without_command=True)
def main(ctx: typer.Context):
    """🧠 Aegis Omniscient v12.0 – run a subcommand, or omit one to analyze the default ticker."""
    if ctx.invoked_subcommand is None:
        cfg = load_config()
        console.print(f"[dim]No command given; analyzing default ticker {cfg['default_ticker']}.[/dim]")
        if not run_single(cfg["default_ticker"]):
            raise typer.Exit(code=1)

@app.command()
def single(
    ticker: str = typer.Argument(None, help="Ticker symbol, e.g. AAPL"),
    export: str = typer.Option(None, "--export", "-e", help="Export format: json, csv, excel, pdf"),
    monte_carlo: bool = typer.Option(False, "--monte-carlo", "-m", help="Run Monte Carlo simulation"),
    no_cache: bool = typer.Option(False, "--no-cache", help="Ignore cached results and refetch"),
    base: str = typer.Option(None, "--base", "-b", help="FCFE base cash flow: auto, ni, fcf, or ocf. Defaults to config 'default_base' (ocf)."),
):
    """Full valuation dashboard for a single ticker."""
    if base is not None and base not in ('auto', 'ni', 'fcf', 'ocf'):
        console.print("[red]--base must be one of: auto, ni, fcf, ocf[/red]")
        raise typer.Exit(code=1)
    if not run_single(ticker, export=export, monte_carlo=monte_carlo, no_cache=no_cache, base_mode=base):
        raise typer.Exit(code=1)

@app.command()
def compare(
    tickers: list[str] = typer.Argument(..., help="Two or more tickers, e.g. AAPL MSFT GOOGL"),
    no_cache: bool = typer.Option(False, "--no-cache", help="Ignore cached results and refetch"),
):
    """Compare key valuation metrics across multiple tickers."""
    if not run_compare(tickers, no_cache=no_cache):
        raise typer.Exit(code=1)

@app.command()
def validate(
    tickers: list[str] = typer.Argument(None, help="Tickers to validate (default: built-in archetype basket)"),
    no_cache: bool = typer.Option(True, "--no-cache/--cache", help="Refetch instead of using cache"),
):
    """Regression harness: value a diverse basket and flag $0 / outlier / garbage results."""
    basket = [t.upper() for t in tickers] if tickers else VALIDATION_BASKET
    ok = run_validate(basket, no_cache=no_cache)
    if not ok:
        raise typer.Exit(code=1)

@app.command()
def add(ticker: str = typer.Argument(..., help="Ticker to add to the watchlist")):
    """Add a ticker to your watchlist."""
    wl = load_watchlist()
    wl.append(ticker.upper())
    wl = save_watchlist(wl)
    console.print(f"[green]✓ Added {ticker.upper()}.[/green] Watchlist: {', '.join(wl)}")

@app.command()
def remove(ticker: str = typer.Argument(..., help="Ticker to remove from the watchlist")):
    """Remove a ticker from your watchlist."""
    wl = load_watchlist()
    if ticker.upper() not in wl:
        console.print(f"[yellow]{ticker.upper()} is not in the watchlist.[/yellow]")
        raise typer.Exit(code=1)
    wl = [t for t in wl if t != ticker.upper()]
    save_watchlist(wl)
    console.print(f"[green]✓ Removed {ticker.upper()}.[/green] Watchlist: {', '.join(wl) or '(empty)'}")

@app.command()
def watchlist(no_cache: bool = typer.Option(False, "--no-cache", help="Ignore cached results and refetch")):
    """Analyze every ticker in your watchlist as a comparison table."""
    wl = load_watchlist()
    if not wl:
        console.print("[yellow]Watchlist is empty. Add tickers with: app.py add TICKER[/yellow]")
        raise typer.Exit(code=1)
    if not run_compare(wl, no_cache=no_cache):
        raise typer.Exit(code=1)

@app.command()
def scenario(
    ticker: str = typer.Argument(..., help="Ticker symbol"),
    growth: float = typer.Option(..., "--growth", "-g", help="Stage-1 FCFE growth rate, e.g. 0.10 for 10%"),
    discount: float = typer.Option(0.09, "--discount", "-d", help="Discount rate (cost of equity)"),
    terminal: float = typer.Option(0.025, "--terminal", "-t", help="Terminal growth rate"),
):
    """Custom 2-stage FCFE DCF: supply your own growth, discount and terminal-growth assumptions."""
    tk = yf.Ticker(ticker)
    try:
        info = tk.info
        cashflow = tk.cashflow
        financials = tk.financials
    except Exception as e:
        console.print(f"[red]Could not fetch data for {ticker.upper()}: {e}[/red]")
        raise typer.Exit(code=1)
    shares = info.get("sharesOutstanding", 0)
    current_price = info.get("currentPrice", info.get("previousClose", 0))
    if not shares:
        console.print(f"[red]Missing share count for {ticker.upper()}.[/red]")
        raise typer.Exit(code=1)
    base_fcfe = normalized_cash_earnings_ps(cashflow, financials, shares)
    val = fcfe_two_stage(base_fcfe, growth, discount, terminal)
    mos = ((val - current_price) / current_price) * 100 if current_price > 0 else 0
    table = Table(title=f"🎯 Scenario FCFE DCF • {ticker.upper()}", box=None)
    table.add_column("Input", style="cyan")
    table.add_column("Value")
    table.add_row("Stage-1 Growth", f"{growth:.1%}")
    table.add_row("Discount (CoE)", f"{discount:.1%}")
    table.add_row("Terminal Growth", f"{terminal:.1%}")
    table.add_row("Base FCFE / Share", f"${base_fcfe:.2f}")
    table.add_row("Current Price", f"${safe_scalar(current_price):.2f}")
    table.add_row("Intrinsic Value", f"[bold gold1]${val:.2f}[/bold gold1]")
    table.add_row("Upside vs Price", f"{mos:.1f}%")
    console.print(table)

@app.command()
def config(
    show: bool = typer.Option(False, "--show", help="Print the current configuration"),
    set_: list[str] = typer.Option(None, "--set", "-s", help="Set a key, e.g. --set default_ticker=MSFT"),
):
    """View or update configuration."""
    cfg = load_config()
    if set_:
        for pair in set_:
            if "=" not in pair:
                console.print(f"[yellow]Ignoring '{pair}' (expected key=value).[/yellow]")
                continue
            key, value = pair.split("=", 1)
            key = key.strip()
            value = value.strip()
            if key not in DEFAULT_CONFIG:
                console.print(f"[yellow]Unknown key '{key}'. Valid keys: {', '.join(DEFAULT_CONFIG)}[/yellow]")
                continue
            # Coerce to the type of the default value.
            default = DEFAULT_CONFIG[key]
            try:
                if isinstance(default, bool):
                    coerced = value.lower() in ("1", "true", "yes", "y", "on")
                elif isinstance(default, int) and not isinstance(default, bool):
                    coerced = int(value)
                elif isinstance(default, float):
                    coerced = float(value)
                else:
                    coerced = value
            except ValueError:
                console.print(f"[red]Could not parse value for {key}: {value}[/red]")
                continue
            cfg[key] = coerced
            console.print(f"[green]✓ {key} = {coerced}[/green]")
        save_config(cfg)
    if show or not set_:
        table = Table(title="⚙️  Configuration", box=None)
        table.add_column("Key", style="cyan")
        table.add_column("Value")
        for k, v in cfg.items():
            table.add_row(k, json.dumps(v) if isinstance(v, dict) else str(v))
        console.print(table)

if __name__ == "__main__":
    app()